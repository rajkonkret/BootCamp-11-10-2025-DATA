from openpyxl import load_workbook

wb1 = load_workbook("tabela_przestawna.xlsx")
wb2 = load_workbook("vgsales_formated.xlsx")

ws1 = wb1.active
ws2 = wb2.active

print(ws1)
print(ws2)

if ws1['A1'].value == ws2['A1'].value:
    print("Wartości A1 są identyczne")
else:
    print(f"Różnica: plik1={ws1['A1'].value}, plik2={ws2["A1"].value}")
# Różnica: plik1=Data, plik2=Rank
