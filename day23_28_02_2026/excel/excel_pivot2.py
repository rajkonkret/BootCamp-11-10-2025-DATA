from openpyxl import Workbook
import openpyxl

data = [
    ["Data", "Kategoria", "Produkt", "Sprzedaj"],
    ["2024-01-01", "Elektronika", "Telefon", 1500],
    ["2024-01-02", "Elektronika", "Telefon", 1500],
    ["2024-01-02", "Odzież", "Koszula", 1500],
    ["2024-01-03", "Odzież", "Kurtka", 1500],
    ["2024-01-04", "Elektronika", "Laptop", 1500],
]
# tworzymy plik excel
wb = Workbook()
ws = wb.active
ws.title = "Dane"

# zapisujemy dane do arkusza
for row in data:
    ws.append(row)

# grupowanie po "Kategoria" i "Produkt
pivot_data = {}
for row in data[1:]:
    category = row[1]
    product = row[2]
    sales = row[3]

    if category not in pivot_data:
        pivot_data[category] = {}

    if product not in pivot_data[category]:
        pivot_data[category][product] = 0

    pivot_data[category][product] += sales

pivot_ws = wb.create_sheet("Tabela Przestawna")

# doajemy nagłówki
pivot_ws.cell(row=1, column=1, value="Kategoria")

product = set(product for category in pivot_data.values() for product in category.keys())
products = sorted(product)  # sortowanie produktów

for col_idx, product in enumerate(products, 2):
    pivot_ws.cell(row=1, column=col_idx, value=product)

# zapisujemy dane tabeli przestawnej
row_idx = 2  # zaczynamy od drugiego wiersza

for category, products_data in pivot_data.items():
    pivot_ws.cell(row=row_idx, column=1, value=category)
    for col_idx, product in enumerate(products, 2):
        sales = products_data.get(product, 0)
        pivot_ws.cell(row=row_idx, column=col_idx, value=sales)

    row_idx += 1

wb.save('tabela_przestawna2.xlsx')
