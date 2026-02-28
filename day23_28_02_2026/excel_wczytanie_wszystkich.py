import glob
from pprint import pprint

from openpyxl import load_workbook

file_list = glob.glob("*.xlsx")  # pokazuje liste plikow o rozszerzeniu .xlsx
print(file_list)
pprint(file_list)
# ['wyniki.xlsx',
#  'xlsxwriter_optimized.xlsx',
#  'edited.xlsx',
#  'dane_zaktualziowane.xlsx',
#  'openpyxl_optimized.xlsx',
#  'tabela_przestawna.xlsx',
#  'tabela_przestawna2.xlsx',
#  'vgsales_formated.xlsx',
#  'xlsxwriter.xlsx',
#  'xlsxwriter5.xlsx']

for file in file_list:
    try:
        wb = load_workbook(file)
        ws = wb.active

        value = ws['A1'].value
        print(f"{file}: A1 = {value}")
    except Exception as e:
        print("Bład:", e)
# wyniki.xlsx: A1 = Imie
# xlsxwriter_optimized.xlsx: A1 = 0
# Bład: File is not a zip file
# dane_zaktualziowane.xlsx: A1 = Rank
# openpyxl_optimized.xlsx: A1 = 0
# tabela_przestawna.xlsx: A1 = Data
# tabela_przestawna2.xlsx: A1 = Data
# vgsales_formated.xlsx: A1 = Rank
# xlsxwriter.xlsx: A1 = None
# xlsxwriter5.xlsx: A1 = Witaj1
