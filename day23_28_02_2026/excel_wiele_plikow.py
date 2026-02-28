from openpyxl import load_workbook

wb1 = load_workbook("tabela_przestawna.xlsx")
wb2 = load_workbook("vgsales_formated.xlsx")

ws1 = wb1.active
ws2 = wb2.active

print(ws1)
print(ws2)
# <Worksheet "Dane">
# <Worksheet "vgsales">

for row in ws1.iter_rows(values_only=True):
    ws2.append(row)

wb2.save("dane_zaktualziowane.xlsx")
